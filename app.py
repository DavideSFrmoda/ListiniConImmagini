import os
import xlrd
import openpyxl
import PIL
from openpyxl.styles import Border, Side


def convert_xls_to_xlsx(file_path):
    assert file_path.endswith('.xls'), 'File must be an .xls file'
    workbook = xlrd.open_workbook(file_path)
    new_workbook = openpyxl.Workbook()
    for sheet_name in workbook.sheet_names():
        worksheet = workbook.sheet_by_name(sheet_name)
        new_worksheet = new_workbook.create_sheet(title=sheet_name)
        for row in range(worksheet.nrows):
            for col in range(worksheet.ncols):
                new_worksheet.cell(row=row+1, column=col +
                                   1).value = worksheet.cell(row, col).value
    new_file_path = file_path.replace('.xls', '.xlsx')
    new_workbook.save(new_file_path)
    return new_file_path


def style_header(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    font = openpyxl.styles.Font(color='FF0000', bold=True, size=12)
    for cell in worksheet[1]:
        cell.font = font

    workbook.save(file_path)

def style_last_row(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    font = openpyxl.styles.Font(color='FF0000', bold=True, size=9)
    for cell in worksheet[worksheet.max_row]:
        cell.font = font

    workbook.save(file_path)

def set_borders(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Create a border style with a thin solid line and black color
    border_style = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))

    last_column = worksheet.max_column
    last_row = worksheet.max_row
    last_cell_str = openpyxl.utils.get_column_letter(last_column) + str(last_row)

    # Apply the border style to a range of cells
    for row in worksheet[f'A1:{last_cell_str}']:
        for cell in row:
            cell.border = border_style

    workbook.save(file_path)


def resize_cells(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(file_path)


def remove_first_sheet(file_path):
    workbook = openpyxl.load_workbook(file_path)
    workbook.remove(workbook.worksheets[0])
    workbook.save(file_path)


def style_image_cells(file_path):
    width = 20
    height = 80

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    worksheet.column_dimensions['A'].width = width
    for i in range(2,worksheet.max_row+1):
        if worksheet.cell(i, 2).value is not None:
            worksheet.row_dimensions[i].height = height
        else:
            break
    workbook.save(file_path)

def add_images(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    folder_path = 'O:\\FRMODA\\trasferite'

    for i in range(2,worksheet.max_row+1):
        cell = worksheet.cell(i, 2)
        sku = cell.value
        if sku is not None:
            try:
                files = [f for f in os.listdir(folder_path) if sku in f]
                assert len(files) > 0, f'No file found for code {sku}'
                file = files[0]
                img_path = os.path.join(folder_path, file)
                assert img_path.endswith('.jpg'), f'File {img_path} is not a jpg image'
                img = openpyxl.drawing.image.Image(img_path)
                size = 80
                img.width = size
                img.height = size
                
                worksheet.add_image(img, f'A{i}')
            except Exception as e:
                print(e)
    workbook.save(file_path)

def set_column_types(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    for col_id in range(4,worksheet.max_column+1):
        for row_id in range(2,worksheet.max_row+1):
            cell = worksheet.cell(row_id, col_id)
            cell.number_format = '0.00'

    workbook.save(file_path)

def style_sheet(path):
    remove_first_sheet(path)
    style_header(path)
    style_last_row(path)
    set_borders(path)
    set_column_types(path)
    resize_cells(path)
    style_image_cells(path)
    add_images(path)


def transform_excel(path):
    path = convert_xls_to_xlsx(path)
    style_sheet(path)
    print(f'cell value: {openpyxl.load_workbook(path).active.cell(1, 1).value}')
    print(
        f'width: {openpyxl.load_workbook(path).active.column_dimensions["A"].width}')
    print(
        f'height: {openpyxl.load_workbook(path).active.row_dimensions[1].height}')